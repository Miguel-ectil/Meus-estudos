import openpyxl
import os
import pyautogui
import subprocess
import pyperclip
import time

# Pegar dados do Excel
arquivo_excel = openpyxl.load_workbook("arquivo_excel.xlsx")
dados = arquivo_excel.active

total_linhas = dados.max_row
total_colunas = dados.max_column

# Caminho onde será salvo o arquivo
desktop_path = os.path.expanduser("~/Documents")
nome_arquivo = "relatorio.txt"
file_path = os.path.join(desktop_path, nome_arquivo)

lista_valores = []  # Lista que será randomizada na calculadora

with open(file_path, "w") as file:
    for linha in range(1, total_linhas + 1):  # para manipular cada linha
        for coluna in range(1, total_colunas + 1):  # para manipular cada coluna
            celula = dados.cell(row=linha, column=coluna)
            valor = celula.value

            if valor is not None:
                conteudo = f"linha {linha}, coluna {coluna}: {valor}\n"
                if type(valor) == int:
                    lista_valores.append(valor)
                file.write(conteudo)  # adiciona as informações do Excel no arquivo

os.system(f"start {file_path}")
time.sleep(2)

# Abre a calculadora do Windows
subprocess.Popen(["calc.exe"])
time.sleep(3)

# Loop para inserir valores e calcular
for i, valor in enumerate(lista_valores):
    pyautogui.write(str(valor))
    if i < len(lista_valores) - 1:
        pyautogui.press('enter')
        pyautogui.press('+')

# Pressiona Enter para obter o resultado final
pyautogui.press('enter')

pyautogui.hotkey('ctrl', 'c')
time.sleep(2)
resultado = pyperclip.paste()
time.sleep(1)
with open(file_path, 'a') as file:  # Use 'a' para modo de anexação
    file.write(resultado)

célula_resultado = dados.cell(row=23, column=2)  # Por exemplo, adicione à primeira linha da última coluna
célula_resultado.value = resultado
time.sleep(2)
# Salve as alterações no arquivo Excel
arquivo_excel.save("arquivo_excel.xlsx")
pyautogui.hotkey('f5')
os.system("start \"\" \"C:\\Program Files\\LibreOffice\\program\\soffice.exe\" --calc arquivo_excel.xlsx")